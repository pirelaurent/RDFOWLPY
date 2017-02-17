# -*- coding: utf-8 -*-
"""
Relit le fichier XML de Ibacs
Crée plusieurs tableurs à partir d'un même nom IBACS_:
IBACS_Entities.xls  qui liste les entités avec leur définition  et leur domaine


"""
import codecs
import sys
import string
import datetime
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Font


myXMLFile="ibacs-mm.xml"
myroot = "IBACS_"

myCytoFile = myroot + "DB.txt"
destExcelFile = myroot + "Entites.xlsx"
destAssocFile = myroot + "Assoc.xlsx"

# pour ne pas multiplier les try catch à chaque fois
def getAttOrNone(someAttributes, aName):
    try :
        value=someAttributes[aName]
    except KeyError:
        value=None
    return value


# MAIN PROG
#------------- pour pouvoir mettre de l'accentué partout
reload(sys)
sys.setdefaultencoding('utf8')
soup = BeautifulSoup(codecs.open(myXMLFile,  "r", "utf-8" ), "xml" )

if False :
    fcyto = codecs.open(myCytoFile, 'w', encoding='utf8')
    # génération du tableau tabulé pour le graphe de dépendance des bases")
    # exploitation des entrée Database
    for database in soup.model.find_all('database'):
        for adomain in database.find_all('domain'):
            dbname = getAttOrNone(database.attrs, "name")
            # db workWith domain
            cible = adomain.text
            fcyto.write("Db\t %s\t workWith \t%s \tdomain \n" % (dbname, cible))
        # recherche et ajout de tous les domaines avec leurs dépendances éventuelles
    for domain in soup.model.find_all('domain'):
        # la balise domain est utilisée aussi au sein des database. On ne veut que les principales
        if domain.parent == soup.model:
            domainName = getAttOrNone(domain.attrs, "name")
            # créer le noeud au cas où pas de dépendance
            fcyto.write("domain \t%s \n" % domainName)
            for depend in domain.find_all('depends-on'):
                cible = depend.text
                fcyto.write("domain \t%s \t depends-on \t%s \tdomain \n" % (domainName, cible))
    fcyto.close
# fin génération graphe de dépendances

# génération du tableur avec les entités par domaine . On remet tout dans le même onglet avec le domaine
if False:
    # create an excel
    wb = Workbook()
    # 1er onglet : les domaines
    ligneDomain = 1
    ws = wb.active
    sheetDomain = "domains"
    ws.title = sheetDomain
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 150
    c = ws['A' + str(ligneDomain)]
    c.value = "Domain"
    c.font = Font(color='FF000000', italic=True, bold=True)
    c = ws['B' + str(ligneDomain)]
    c.value = "Description"
    c.font = Font(color='FF000000', italic=True, bold=True)

    # 2eme onglet : les entités
    ligneEntity = 1
    sheetEntities = "entities"
    wb.create_sheet(sheetEntities)
    ws = wb.get_sheet_by_name(sheetEntities)
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 120
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['F'].width = 20
    c = ws['A' + str(ligneEntity)]
    c.value = "Entity"
    c.font = Font(color='FF000000', italic=True, bold=True)
    c = ws['B' + str(ligneEntity)]
    c.value = "Description"
    c.font = Font(color='FF000000', italic=True, bold=True)
    c = ws['C' + str(ligneEntity)]
    c.value = "Attributs"
    c.font = Font(color='FF000000', italic=True, bold=True)
    c = ws['D' + str(ligneEntity)]
    c.value = "Implements"
    c.font = Font(color='FF000000', italic=True, bold=True)
    c = ws['E' + str(ligneEntity)]
    c.value = "associations"
    c.font = Font(color='FF000000', italic=True, bold=True)
    c = ws['F' + str(ligneEntity)]
    c.value = "Domain"
    c.font = Font(color='FF000000', italic=True, bold=True)

    # Balayage des domaines

    for domain in soup.model.find_all('domain'):
        if domain.parent == soup.model :
            # dans le premier onglet, on note le domain et sa description
            ws = wb.get_sheet_by_name(sheetDomain)
            ligneDomain += 1
            domainName = getAttOrNone(domain.attrs,"name")
            adesc = domain.find("description")
            if adesc is not None:
                langue = getAttOrNone(adesc.attrs, "lang")
                description = adesc.text
            else :
                description = "*** aucune description ***"
                langue = ""
            ws['A' + str(ligneDomain)] = domainName
            ws['B' + str(ligneDomain)] = description
            # 2ème onglet : Entités
            for entity in  domain.find_all ('element'):
                ws = wb.get_sheet_by_name(sheetEntities)
                ligneEntity += 1
                atype = getAttOrNone(entity.attrs,"type")
                # on nettoie le préfixe java
                point = atype.rindex('.')
                atype = atype [point+1:]
                adesc = entity.find('description')
                if adesc is not None:
                    langue = getAttOrNone(adesc.attrs, "lang")
                    description = adesc.text
                else:
                    description = "*** aucune description ***"
                    langue = ""
                ws['A' + str(ligneEntity)] = atype
                ws['B' + str(ligneEntity)] = description
                # on compte les attributs pour la complexité
                # Attributs
                nbre = len (entity.find_all('property'))
                nbre += len (entity.find_all('id'))
                ws['C' + str(ligneEntity)] = nbre
                # Interfaces
                nbreItf = len (entity.find_all('implements'))
                ws['D' + str(ligneEntity)] = nbreItf
                # relation
                nbreTip= len (entity.find_all('tip'))
                ws['E' + str(ligneEntity)] = nbreTip
                # domain d'origine
                ws['F' + str(ligneEntity)] = domainName
    try:
        wb.save(destExcelFile)
    except IOError as e:
        print "*** impossible de créer le fichier excel (penser à le fermer): "
        print e

# -------------------------------------------------------------------------
# tableur pour repérer les associations entre entités

if True:
    dictEntityDomain = {}
    # create an excel
    wb = Workbook()
    ligneTableur = 1
    ws = wb.active
    sheetGraphe = "Graphe"
    ws.title = sheetGraphe
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 30
    ws.column_dimensions['F'].width = 30
    ws.column_dimensions['G'].width = 30

    c = ws['A' + str(ligneTableur)]
    c.value = "Entity"
    c.font = Font(color='FF000000', italic=True, bold=True)
    c = ws['B' + str(ligneTableur)]
    c.value = "Complexity"
    c.font = Font(color='FF000000', italic=True, bold=True)
    c = ws['C' + str(ligneTableur)]
    c.value = "typeAssoc"
    c.font = Font(color='FF000000', italic=True, bold=True)
    c = ws['D' + str(ligneTableur)]
    c.value = "destination"
    c.font = Font(color='FF000000', italic=True, bold=True)
    c = ws['E' + str(ligneTableur)]
    c.value = "domain"
    c.font = Font(color='FF000000', italic=True, bold=True)
    c = ws['F' + str(ligneTableur)]
    c.value = "dest. domain"
    c.font = Font(color='FF000000', italic=True, bold=True)
    c = ws['G' + str(ligneTableur)]
    c.value = "cross"
    c.font = Font(color='FF000000', italic=True, bold=True)
    # Balayage des domaines

    for domain in soup.model.find_all('domain'):
        if domain.parent == soup.model:
            domainName = getAttOrNone(domain.attrs, "name")
            # les  Entités
            for entity in domain.find_all('element'):
                ligneTableur += 1
                atype = getAttOrNone(entity.attrs, "type")
                # on nettoie le préfixe java
                point = atype.rindex('.')
                atype = atype[point + 1:]
                # on note déja entité et domaine
                ws['A' + str(ligneTableur)] = atype
                ws['E' + str(ligneTableur)] = domainName
                dictEntityDomain[atype] = domainName
                # on compte les attributs + interfaces pour la complexité
                # Attributs
                nbre = len(entity.find_all('property'))
                nbre += len(entity.find_all('id'))
                nbre += len(entity.find_all('implements'))
                ws['B' + str(ligneTableur)] = nbre
                # relation
                for atip in entity.find_all('tip'):
                    typetip = getAttOrNone(atip.attrs, "type")
                    # garder la partie droite uniquement
                    point = typetip.rindex('.')
                    typetip = typetip[point + 1:]
                    multiTip = getAttOrNone(atip.attrs, "multiplicity")
                    ligneTableur += 1
                    ws['A' + str(ligneTableur)] = atype
                    ws['C' + str(ligneTableur)] = multiTip
                    ws['D' + str(ligneTableur)] = typetip
                    ws['E' + str(ligneTableur)] = domainName
                    # après on refait une passe pour les destinations

    for row in range(2, ligneTableur + 1):
        dest = ws['D' + str(row)].value
        if dest is not None:
            sourceDomain = ws['E' + str(row)].value
            # on recherhe quel est le domain

            if dictEntityDomain.has_key(dest):
                destDomain = dictEntityDomain[dest]

                ws['F' + str(row)] = destDomain
                if sourceDomain != destDomain:
                    ws['G' + str(row)] = "CrossDomain"
                else:
                    ws['G' + str(row)] = "sameDomain"
        else:
            pass

    try:
        wb.save(destAssocFile)
    except IOError as e:
        print "*** impossible de créer le fichier excel (penser à le fermer): "
        print e
